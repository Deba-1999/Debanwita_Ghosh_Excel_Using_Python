import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border,Side

#create Data Frame
Name= ['Debanwita','Akansha','Raja','Jitu','Trisha','Bishu']
Emp_Id=['sky002','sky021','sky009','sky010','sky025','sky034']
Email=['dgsky002@gmail.com','apsky021@gmail.com','rcsky009@gmail.com','jssky010@gmail.com','tdsky025@gmail.com','bdsky034@gmail.com']
Department=['Accounts','Sales','HR','Sales','Data Analyst','QA']
Salary=[25000,30000,50000,32000,30000,35000]
columns=['Name','Emp_Id','Email','Department','Salary']
df=pd.DataFrame(list(zip(Name,Emp_Id,Email,Department,Salary)),columns=columns)
print(df)

#Data Frame to Excel
df.to_excel('Excel_Using_Python.xlsx')
df.to_excel('Excel_Using_Python.xlsx',sheet_name='Project')
wb= load_workbook('Excel_Using_Python.xlsx')
ws=wb.active

#insert a row in the top , put there location
ws.insert_rows(1)
ws.cell(row=1,column=3).value='Office Dataset'
wb.save('Excel_Using_Python.xlsx')

#merge cell
ws.merge_cells('c1:d1')
wb.save('Excel_Using_Python.xlsx')

#entry the cell's name
ws.cell(row=9,column=5).value='SUM :'
ws.cell(row=10,column=5).value='Average :'
wb.save('Excel_Using_Python.xlsx')

#font designing
ws.cell(row=1,column=3).font=Font(b=True,i=True)
ws.cell(row=9,column=5).font=Font(b=True)
ws.cell(row=10,column=5).font=Font(b=True)
wb.save('Excel_Using_Python.xlsx')

#Change the Sheet Name
sheet=wb.get_sheet_by_name('Project')
wb.save('Excel_Using_Python.xlsx')

#apply formula for getting result
v1=sheet['f3'].value
v2=sheet['f4'].value
v3=sheet['f5'].value
v4=sheet['f6'].value
v5=sheet['f7'].value
v6=sheet['f8'].value
sheet['f9']=v1+v2+v3+v4+v5+v6
value=v1+v2+v3+v4+v5+v6
sheet['f10']=(value/6)
wb.save('Excel_Using_Python.xlsx')

#change cells colour
fill_color1=PatternFill(patternType='solid',fgColor='DFEF09')
ws.cell(row=1,column=3).fill=fill_color1
fill_color2=PatternFill(patternType='solid',fgColor='4AE2F7')
ws.cell(row=9,column=5).fill=fill_color2
fill_color3=PatternFill(patternType='solid',fgColor='4AE2F7')
ws.cell(row=10,column=5).fill=fill_color3
fill_color4=PatternFill(patternType='solid',fgColor='55D2FA')
ws.cell(row=3,column=1).fill=fill_color4
ws.cell(row=4,column=1).fill=fill_color4
ws.cell(row=5,column=1).fill=fill_color4
ws.cell(row=6,column=1).fill=fill_color4
ws.cell(row=7,column=1).fill=fill_color4
ws.cell(row=8,column=1).fill=fill_color4

fill_color5=PatternFill(patternType='solid',fgColor='EF4BD1')
ws.cell(row=2,column=2).fill=fill_color5
ws.cell(row=2,column=3).fill=fill_color5
ws.cell(row=2,column=4).fill=fill_color5
ws.cell(row=2,column=5).fill=fill_color5
ws.cell(row=2,column=6).fill=fill_color5
wb.save('Excel_Using_Python.xlsx')

#adding border
ws['E9'].border= Border (top= Side(style='thick',color='020D78'),right= Side(style='thick',color='020D78'),left= Side(style='thick',color='020D78'),bottom= Side(style='thick',color='020D78'))
ws['E10'].border= Border (top= Side(style='thick',color='020D78'),right= Side(style='thick',color='020D78'),left= Side(style='thick',color='020D78'),bottom= Side(style='thick',color='020D78'))
wb.save('Excel_Using_Python.xlsx')

from openpyxl.chart import BarChart,Reference
chart= BarChart()
data= Reference(ws,min_col=6,min_row=3,max_col=6,max_row=8)
chart.title='Employee'
chart.y_axis.title = "Salary"
chart.add_data(data,titles_from_data=True)
ws.add_chart(chart,"h2")
wb.save('Excel_Using_Python.xlsx')